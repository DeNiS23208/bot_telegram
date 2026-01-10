#!/usr/bin/env python3
"""
Скрипт для генерации Excel отчета из базы данных бота
Создает красивый и понятный отчет для админа
"""
import sqlite3
import os
import sys
from datetime import datetime, timezone, timedelta
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from config import PAYMENT_AMOUNT_RUB, BONUS_WEEK_PRICE_RUB, PRODUCTION_PRICE_RUB, BONUS_WEEK_START_DATE, BONUS_WEEK_END_DATE

# Московское время
MoscowTz = pytz.timezone('Europe/Moscow')

# Путь к базе данных
DB_PATH = os.getenv("DB_PATH", "bot.db")

# Цвета для оформления
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_datetime(dt_str):
    """Форматирует дату в читаемый вид в МСК времени"""
    if not dt_str:
        return "—"
    try:
        dt = datetime.fromisoformat(dt_str.replace('Z', '+00:00'))
        # Если нет timezone, считаем что UTC
        if dt.tzinfo is None:
            dt = pytz.utc.localize(dt)
        # Конвертируем в МСК
        moscow_dt = dt.astimezone(MoscowTz)
        return moscow_dt.strftime("%d.%m.%Y %H:%M:%S МСК")
    except:
        return dt_str

def format_time_remaining(expires_at_str, now):
    """Форматирует оставшееся время до окончания подписки"""
    if not expires_at_str:
        return "—"
    try:
        expires_dt = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
        if expires_dt.tzinfo is None:
            expires_dt = pytz.utc.localize(expires_dt)
        
        if expires_dt <= now:
            return "Истекла"
        
        delta = expires_dt - now
        days = delta.days
        hours = delta.seconds // 3600
        minutes = (delta.seconds % 3600) // 60
        
        if days > 0:
            return f"{days} дн. {hours} ч."
        elif hours > 0:
            return f"{hours} ч. {minutes} мин."
        else:
            return f"{minutes} мин."
    except:
        return "—"

def format_duration_days(starts_at_str, expires_at_str):
    """Вычисляет длительность подписки в днях"""
    if not starts_at_str or not expires_at_str:
        return "—"
    try:
        starts_dt = datetime.fromisoformat(starts_at_str.replace('Z', '+00:00'))
        expires_dt = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
        if starts_dt.tzinfo is None:
            starts_dt = pytz.utc.localize(starts_dt)
        if expires_dt.tzinfo is None:
            expires_dt = pytz.utc.localize(expires_dt)
        
        delta = expires_dt - starts_dt
        days = delta.days
        hours = delta.seconds // 3600
        if hours >= 12:
            days += 1
        return f"{days} дн."
    except:
        return "—"

def format_status(status):
    """Форматирует статус платежа с цветом"""
    status_map = {
        "succeeded": "✅ Успешно",
        "pending": "⏳ Ожидает",
        "canceled": "❌ Отменен",
        "expired": "⏰ Истек"
    }
    return status_map.get(status, status)

def format_gender(gender):
    """Форматирует пол"""
    gender_map = {
        "male": "Мужской",
        "female": "Женский",
        "other": "Другое"
    }
    return gender_map.get(gender, gender or "—")

def is_bonus_week_payment(created_at_str):
    """Проверяет, был ли платеж во время бонусной недели"""
    if not created_at_str:
        return False
    try:
        payment_dt = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
        if payment_dt.tzinfo is None:
            payment_dt = pytz.utc.localize(payment_dt)
        
        return BONUS_WEEK_START_DATE <= payment_dt <= BONUS_WEEK_END_DATE
    except:
        return False

def get_payment_amount(created_at_str):
    """Возвращает реальную сумму платежа"""
    if is_bonus_week_payment(created_at_str):
        return BONUS_WEEK_PRICE_RUB
    return PRODUCTION_PRICE_RUB

def format_auto_renewal_status(auto_renewal_enabled, attempts, last_attempt_at):
    """Форматирует статус автопродления"""
    if not auto_renewal_enabled:
        return "❌ Не включено"
    if attempts is None or attempts == 0:
        return "✅ Включено"
    if attempts >= 3:
        return f"❌ Отключено (3 попытки)"
    return f"⚠️ Попытки ({attempts}/3)"

def create_users_sheet(wb, conn):
    """Создает лист с пользователями"""
    ws = wb.active
    ws.title = "Пользователи"
    
    # Заголовок
    ws['A1'] = "СПИСОК ПОЛЬЗОВАТЕЛЕЙ"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:P1')
    
    # Заголовки колонок
    headers = [
        "ID Telegram", "Username", "Дата регистрации", "Количество платежей",
        "Доступ с", "Доступ до", "Осталось времени", "Длительность (дн.)",
        "Статус платежа", "Дата последнего платежа", "Статус на канале", "Добавлен/Забанен",
        "Форма заполнена", "Дата заполнения формы", "Попытки автопродления", "Последняя попытка"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Получаем данные с дополнительной информацией
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            u.telegram_id, 
            u.username, 
            u.created_at,
            COUNT(DISTINCT p.id) as payment_count,
            s.starts_at,
            s.expires_at,
            (SELECT status FROM payments WHERE telegram_id = u.telegram_id ORDER BY id DESC LIMIT 1) as last_payment_status,
            (SELECT MAX(created_at) FROM payments WHERE telegram_id = u.telegram_id) as last_payment_date,
            au.approved_at,
            (SELECT revoked FROM invite_links WHERE telegram_user_id = u.telegram_id ORDER BY created_at DESC LIMIT 1) as last_link_revoked,
            u.form_filled,
            u.form_filled_at,
            s.auto_renewal_attempts,
            s.last_auto_renewal_attempt_at
        FROM users u
        LEFT JOIN payments p ON u.telegram_id = p.telegram_id
        LEFT JOIN subscriptions s ON u.telegram_id = s.telegram_id
        LEFT JOIN approved_users au ON u.telegram_id = au.telegram_user_id
        GROUP BY u.telegram_id
        ORDER BY u.created_at DESC
    """)
    
    row = 4
    now = datetime.now(timezone.utc)
    for record in cur.fetchall():
        telegram_id, username, created_at, payment_count, starts_at, expires_at, last_payment_status, last_payment_date, approved_at, last_link_revoked, form_filled, form_filled_at, auto_renewal_attempts, last_auto_renewal_attempt_at = record
        
        ws.cell(row=row, column=1, value=telegram_id).border = BORDER
        ws.cell(row=row, column=2, value=username or "—").border = BORDER
        ws.cell(row=row, column=3, value=format_datetime(created_at)).border = BORDER
        ws.cell(row=row, column=4, value=payment_count).border = BORDER
        
        # Период доступа
        ws.cell(row=row, column=5, value=format_datetime(starts_at)).border = BORDER
        expires_cell = ws.cell(row=row, column=6, value=format_datetime(expires_at))
        expires_cell.border = BORDER
        
        # Оставшееся время
        time_remaining = format_time_remaining(expires_at, now)
        remaining_cell = ws.cell(row=row, column=7, value=time_remaining)
        remaining_cell.border = BORDER
        if time_remaining != "—" and time_remaining != "Истекла":
            remaining_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Длительность подписки
        duration = format_duration_days(starts_at, expires_at)
        ws.cell(row=row, column=8, value=duration).border = BORDER
        
        # Статус платежа
        status_cell = ws.cell(row=row, column=9, value=format_status(last_payment_status) if last_payment_status else "—")
        status_cell.border = BORDER
        if last_payment_status == "succeeded":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif last_payment_status == "pending":
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif last_payment_status in ["canceled", "expired"]:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Дата последнего платежа
        ws.cell(row=row, column=10, value=format_datetime(last_payment_date)).border = BORDER
        
        # Статус на канале (детальный)
        channel_status = "—"
        channel_simple_status = "—"
        if approved_at:
            try:
                expires_dt = datetime.fromisoformat(expires_at.replace('Z', '+00:00')) if expires_at else None
                if expires_dt and expires_dt.tzinfo is None:
                    expires_dt = pytz.utc.localize(expires_dt)
                
                if expires_dt and expires_dt > now:
                    # Подписка активна
                    channel_status = f"✅ Добавлен: {format_datetime(approved_at)}"
                    channel_simple_status = "✅ Добавлен"
                elif last_link_revoked and expires_dt:
                    # Забанен (ссылка отозвана и подписка истекла)
                    channel_status = f"❌ Забанен (примерно): {format_datetime(expires_at)}"
                    channel_simple_status = "❌ Забанен"
                else:
                    # Был добавлен, но статус неясен
                    channel_status = f"ℹ️ Добавлен: {format_datetime(approved_at)}"
                    channel_simple_status = "✅ Добавлен"
            except:
                channel_status = f"ℹ️ Добавлен: {format_datetime(approved_at)}"
                channel_simple_status = "✅ Добавлен"
        
        channel_status_cell = ws.cell(row=row, column=11, value=channel_status)
        channel_status_cell.border = BORDER
        if "✅" in channel_status:
            channel_status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif "❌" in channel_status:
            channel_status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Простой статус на канале (Добавлен/Забанен)
        simple_status_cell = ws.cell(row=row, column=12, value=channel_simple_status)
        simple_status_cell.border = BORDER
        simple_status_cell.alignment = Alignment(horizontal='center', vertical='center')
        if channel_simple_status == "✅ Добавлен":
            simple_status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            simple_status_cell.font = Font(bold=True)
        elif channel_simple_status == "❌ Забанен":
            simple_status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            simple_status_cell.font = Font(bold=True)
        
        # Форма заполнена
        form_filled_text = "Да" if form_filled else "Нет"
        form_cell = ws.cell(row=row, column=13, value=form_filled_text)
        form_cell.border = BORDER
        if form_filled:
            form_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Дата заполнения формы
        ws.cell(row=row, column=14, value=format_datetime(form_filled_at)).border = BORDER
        
        # Попытки автопродления
        attempts_text = str(auto_renewal_attempts) if auto_renewal_attempts else "0"
        attempts_cell = ws.cell(row=row, column=15, value=attempts_text)
        attempts_cell.border = BORDER
        if auto_renewal_attempts and auto_renewal_attempts >= 3:
            attempts_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif auto_renewal_attempts and auto_renewal_attempts > 0:
            attempts_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Дата последней попытки автопродления
        ws.cell(row=row, column=16, value=format_datetime(last_auto_renewal_attempt_at)).border = BORDER
        
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 35
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 18
    ws.column_dimensions['N'].width = 22
    ws.column_dimensions['O'].width = 22
    ws.column_dimensions['P'].width = 25

def create_payments_sheet(wb, conn):
    """Создает лист с платежами"""
    ws = wb.create_sheet("Платежи")
    
    # Заголовок
    ws['A1'] = "ИСТОРИЯ ПЛАТЕЖЕЙ"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:M1')
    
    # Заголовки колонок
    headers = [
        "ID Платежа", "ID Пользователя", "Username", "Статус", 
        "Сумма (руб)", "Реальная сумма", "Тип платежа", "Бонусная неделя",
        "Дата создания", "Дата обработки", "Время обработки", "Ссылка создана", "Метод оплаты"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Получаем данные
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            p.payment_id,
            p.telegram_id,
            u.username,
            p.status,
            p.created_at,
            pp.processed_at,
            il.created_at as link_created
        FROM payments p
        LEFT JOIN users u ON p.telegram_id = u.telegram_id
        LEFT JOIN processed_payments pp ON p.payment_id = pp.payment_id
        LEFT JOIN invite_links il ON p.payment_id = il.payment_id
        ORDER BY p.created_at DESC
    """)
    
    row = 4
    for record in cur.fetchall():
        payment_id, telegram_id, username, status, created_at, processed_at, link_created = record
        
        ws.cell(row=row, column=1, value=payment_id).border = BORDER
        ws.cell(row=row, column=2, value=telegram_id).border = BORDER
        ws.cell(row=row, column=3, value=username or "—").border = BORDER
        status_cell = ws.cell(row=row, column=4, value=format_status(status))
        status_cell.border = BORDER
        
        # Определяем тип платежа и сумму
        is_bonus = is_bonus_week_payment(created_at)
        real_amount = get_payment_amount(created_at)
        payment_type = "Бонусная неделя" if is_bonus else "Продакшн"
        
        # Если есть saved_payment_method_id - это автопродление
        cur2 = conn.cursor()
        cur2.execute("""
            SELECT saved_payment_method_id FROM subscriptions WHERE telegram_id = ?
        """, (telegram_id,))
        sub_row = cur2.fetchone()
        if sub_row and sub_row[0]:
            # Проверяем, был ли это автоплатеж (по времени создания относительно подписки)
            cur2.execute("""
                SELECT starts_at FROM subscriptions WHERE telegram_id = ?
            """, (telegram_id,))
            sub_start = cur2.fetchone()
            if sub_start and sub_start[0]:
                try:
                    sub_start_dt = datetime.fromisoformat(sub_start[0].replace('Z', '+00:00'))
                    payment_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    if payment_dt.tzinfo is None:
                        payment_dt = pytz.utc.localize(payment_dt)
                    if sub_start_dt.tzinfo is None:
                        sub_start_dt = pytz.utc.localize(sub_start_dt)
                    
                    # Если платеж создан после начала подписки - это автопродление
                    if payment_dt > sub_start_dt + timedelta(minutes=5):
                        payment_type = "Автопродление"
                except:
                    pass
        
        ws.cell(row=row, column=5, value=PAYMENT_AMOUNT_RUB).border = BORDER  # Базовая сумма из конфига
        ws.cell(row=row, column=6, value=real_amount).border = BORDER  # Реальная сумма
        ws.cell(row=row, column=7, value=payment_type).border = BORDER
        bonus_cell = ws.cell(row=row, column=8, value="Да" if is_bonus else "Нет")
        bonus_cell.border = BORDER
        if is_bonus:
            bonus_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws.cell(row=row, column=9, value=format_datetime(created_at)).border = BORDER
        ws.cell(row=row, column=10, value=format_datetime(processed_at)).border = BORDER
        
        # Время обработки (разница между processed_at и created_at)
        processing_time = "—"
        if processed_at and created_at:
            try:
                created_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                processed_dt = datetime.fromisoformat(processed_at.replace('Z', '+00:00'))
                if created_dt.tzinfo is None:
                    created_dt = pytz.utc.localize(created_dt)
                if processed_dt.tzinfo is None:
                    processed_dt = pytz.utc.localize(processed_dt)
                
                delta = processed_dt - created_dt
                minutes = int(delta.total_seconds() / 60)
                if minutes < 60:
                    processing_time = f"{minutes} мин."
                else:
                    hours = minutes // 60
                    mins = minutes % 60
                    processing_time = f"{hours} ч. {mins} мин."
            except:
                pass
        
        ws.cell(row=row, column=11, value=processing_time).border = BORDER
        ws.cell(row=row, column=12, value="Да" if link_created else "Нет").border = BORDER
        
        # Метод оплаты (из payment_id можно определить, но обычно не хранится)
        # Пока оставляем пустым или определяем по паттерну
        payment_method = "—"
        if payment_id:
            # Можно попробовать определить по паттерну ID
            if "sber" in payment_id.lower() or "sbp" in payment_id.lower():
                payment_method = "СБП/SberPay"
            else:
                payment_method = "Карта"
        ws.cell(row=row, column=13, value=payment_method).border = BORDER
        
        # Цветовая индикация статуса
        if status == "succeeded":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif status == "pending":
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif status in ["canceled", "expired"]:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 15

def create_subscriptions_sheet(wb, conn):
    """Создает лист с подписками"""
    ws = wb.create_sheet("Подписки")
    
    # Заголовок
    ws['A1'] = "АКТИВНЫЕ И ИСТЕКШИЕ ПОДПИСКИ"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:O1')
    
    # Заголовки колонок
    headers = [
        "ID Пользователя", "Username", "Начало доступа", "Окончание доступа", 
        "Осталось времени", "Длительность (дн.)", "Статус", "Автопродление",
        "Попытки автопродления", "Последняя попытка", "Статус автопродления",
        "Сохранена карта", "ID способа оплаты", "Уведомление отправлено"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Получаем данные
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            s.telegram_id,
            u.username,
            s.starts_at,
            s.expires_at,
            s.auto_renewal_enabled,
            s.saved_payment_method_id,
            s.subscription_expired_notified,
            s.auto_renewal_attempts,
            s.last_auto_renewal_attempt_at
        FROM subscriptions s
        LEFT JOIN users u ON s.telegram_id = u.telegram_id
        ORDER BY s.expires_at DESC
    """)
    
    row = 4
    now = datetime.now(timezone.utc)
    for record in cur.fetchall():
        telegram_id, username, starts_at, expires_at, auto_renewal, saved_card, notified, attempts, last_attempt_at = record
        ws.cell(row=row, column=1, value=telegram_id).border = BORDER
        ws.cell(row=row, column=2, value=username or "—").border = BORDER
        ws.cell(row=row, column=3, value=format_datetime(starts_at)).border = BORDER
        expires_cell = ws.cell(row=row, column=4, value=format_datetime(expires_at))
        expires_cell.border = BORDER
        
        # Оставшееся время
        time_remaining = format_time_remaining(expires_at, now)
        remaining_cell = ws.cell(row=row, column=5, value=time_remaining)
        remaining_cell.border = BORDER
        if time_remaining != "—" and time_remaining != "Истекла":
            remaining_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        
        # Длительность подписки
        duration = format_duration_days(starts_at, expires_at)
        ws.cell(row=row, column=6, value=duration).border = BORDER
        
        # Определяем статус
        try:
            expires_dt = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
            if expires_dt.tzinfo is None:
                expires_dt = pytz.utc.localize(expires_dt)
            if expires_dt > now:
                status = "✅ Активна"
                status_cell = ws.cell(row=row, column=7, value=status)
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status = "⏰ Истекла"
                status_cell = ws.cell(row=row, column=7, value=status)
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        except:
            status_cell = ws.cell(row=row, column=7, value="—")
        status_cell.border = BORDER
        
        ws.cell(row=row, column=8, value="Да" if auto_renewal else "Нет").border = BORDER
        
        # Попытки автопродления
        attempts_text = str(attempts) if attempts else "0"
        attempts_cell = ws.cell(row=row, column=9, value=attempts_text)
        attempts_cell.border = BORDER
        if attempts and attempts >= 3:
            attempts_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif attempts and attempts > 0:
            attempts_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Дата последней попытки
        ws.cell(row=row, column=10, value=format_datetime(last_attempt_at)).border = BORDER
        
        # Статус автопродления (детальный)
        auto_status = format_auto_renewal_status(auto_renewal, attempts, last_attempt_at)
        auto_status_cell = ws.cell(row=row, column=11, value=auto_status)
        auto_status_cell.border = BORDER
        if "✅" in auto_status:
            auto_status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif "❌" in auto_status:
            auto_status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif "⚠️" in auto_status:
            auto_status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws.cell(row=row, column=12, value="Да" if saved_card else "Нет").border = BORDER
        
        # ID способа оплаты (первые и последние символы для безопасности)
        payment_method_id_display = "—"
        if saved_card:
            if len(saved_card) > 8:
                payment_method_id_display = f"{saved_card[:4]}...{saved_card[-4:]}"
            else:
                payment_method_id_display = "***"
        ws.cell(row=row, column=13, value=payment_method_id_display).border = BORDER
        
        ws.cell(row=row, column=14, value="Да" if notified else "Нет").border = BORDER
        
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20

def create_invite_links_sheet(wb, conn):
    """Создает лист с пригласительными ссылками"""
    ws = wb.create_sheet("Пригласительные ссылки")
    
    # Заголовок
    ws['A1'] = "ИСТОРИЯ ПРИГЛАСИТЕЛЬНЫХ ССЫЛОК"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    # Заголовки колонок
    headers = [
        "Ссылка", "ID Пользователя", "Username", "ID Платежа", 
        "Дата создания", "Статус"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Получаем данные
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            il.invite_link,
            il.telegram_user_id,
            u.username,
            il.payment_id,
            il.created_at,
            il.revoked
        FROM invite_links il
        LEFT JOIN users u ON il.telegram_user_id = u.telegram_id
        ORDER BY il.created_at DESC
    """)
    
    row = 4
    for record in cur.fetchall():
        invite_link, telegram_id, username, payment_id, created_at, revoked = record
        ws.cell(row=row, column=1, value=invite_link).border = BORDER
        ws.cell(row=row, column=2, value=telegram_id).border = BORDER
        ws.cell(row=row, column=3, value=username or "—").border = BORDER
        ws.cell(row=row, column=4, value=payment_id).border = BORDER
        ws.cell(row=row, column=5, value=format_datetime(created_at)).border = BORDER
        status_cell = ws.cell(row=row, column=6, value="Отозвана" if revoked else "Активна")
        status_cell.border = BORDER
        if revoked:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

def create_forms_sheet(wb, conn):
    """Создает лист с заполненными формами"""
    ws = wb.create_sheet("Заполненные формы")
    
    # Заголовок
    ws['A1'] = "ДАННЫЕ ЗАПОЛНЕННЫХ ФОРМ"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:L1')
    
    # Заголовки колонок
    headers = [
        "ID Пользователя", "Username", "Дата заполнения", "Имя",
        "Телефон", "Email", "Город", "Пол", "Направление деятельности",
        "Согласие на ПД", "Согласие с Офертой"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Получаем данные
    cur = conn.cursor()
    # Проверяем, существует ли таблица
    cur.execute("""
        SELECT name FROM sqlite_master WHERE type='table' AND name='daily_form_submissions'
    """)
    table_exists = cur.fetchone()
    
    if table_exists:
        cur.execute("""
            SELECT 
                telegram_id,
                username,
                submitted_at,
                name,
                phone,
                email,
                city,
                gender,
                activity,
                privacy_accepted,
                offer_accepted
            FROM daily_form_submissions
            ORDER BY submitted_at DESC
        """)
        
        row = 4
        for record in cur.fetchall():
            telegram_id, username, submitted_at, name, phone, email, city, gender, activity, privacy_accepted, offer_accepted = record
            
            ws.cell(row=row, column=1, value=telegram_id).border = BORDER
            ws.cell(row=row, column=2, value=username or "—").border = BORDER
            ws.cell(row=row, column=3, value=format_datetime(submitted_at)).border = BORDER
            ws.cell(row=row, column=4, value=name or "—").border = BORDER
            ws.cell(row=row, column=5, value=phone or "—").border = BORDER
            ws.cell(row=row, column=6, value=email or "—").border = BORDER
            ws.cell(row=row, column=7, value=city or "—").border = BORDER
            ws.cell(row=row, column=8, value=format_gender(gender)).border = BORDER
            ws.cell(row=row, column=9, value=activity or "—").border = BORDER
            ws.cell(row=row, column=10, value="Да" if privacy_accepted else "Нет").border = BORDER
            ws.cell(row=row, column=11, value="Да" if offer_accepted else "Нет").border = BORDER
            
            row += 1
    else:
        # Таблица не существует - добавляем сообщение
        ws.cell(row=4, column=1, value="Таблица daily_form_submissions не найдена в базе данных")
        ws.merge_cells('A4:L4')
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18

def create_auto_renewal_sheet(wb, conn):
    """Создает лист с детальной информацией об автопродлении"""
    ws = wb.create_sheet("Автопродление")
    
    # Заголовок
    ws['A1'] = "ДЕТАЛЬНАЯ СТАТИСТИКА ПО АВТОПРОДЛЕНИЮ"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')
    
    # Заголовки колонок
    headers = [
        "ID Пользователя", "Username", "Автопродление включено",
        "Попытки (X/3)", "Дата последней попытки", "Статус сохраненной карты",
        "ID способа оплаты", "Дата следующей попытки"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    
    # Получаем данные
    cur = conn.cursor()
    cur.execute("""
        SELECT 
            s.telegram_id,
            u.username,
            s.auto_renewal_enabled,
            s.auto_renewal_attempts,
            s.last_auto_renewal_attempt_at,
            s.saved_payment_method_id,
            s.expires_at
        FROM subscriptions s
        LEFT JOIN users u ON s.telegram_id = u.telegram_id
        WHERE s.auto_renewal_enabled = 1 OR s.auto_renewal_attempts > 0
        ORDER BY s.auto_renewal_attempts DESC, s.last_auto_renewal_attempt_at DESC
    """)
    
    row = 4
    now = datetime.now(timezone.utc)
    for record in cur.fetchall():
        telegram_id, username, auto_renewal_enabled, attempts, last_attempt_at, saved_payment_method_id, expires_at = record
        
        ws.cell(row=row, column=1, value=telegram_id).border = BORDER
        ws.cell(row=row, column=2, value=username or "—").border = BORDER
        
        auto_enabled_text = "Да" if auto_renewal_enabled else "Нет"
        auto_cell = ws.cell(row=row, column=3, value=auto_enabled_text)
        auto_cell.border = BORDER
        if auto_renewal_enabled:
            auto_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            auto_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        attempts_text = f"{attempts or 0}/3"
        attempts_cell = ws.cell(row=row, column=4, value=attempts_text)
        attempts_cell.border = BORDER
        if attempts and attempts >= 3:
            attempts_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif attempts and attempts > 0:
            attempts_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws.cell(row=row, column=5, value=format_datetime(last_attempt_at)).border = BORDER
        
        card_status = "Да" if saved_payment_method_id else "Нет"
        card_cell = ws.cell(row=row, column=6, value=card_status)
        card_cell.border = BORDER
        if saved_payment_method_id:
            card_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # ID способа оплаты (частично)
        payment_method_id_display = "—"
        if saved_payment_method_id:
            if len(saved_payment_method_id) > 8:
                payment_method_id_display = f"{saved_payment_method_id[:4]}...{saved_payment_method_id[-4:]}"
            else:
                payment_method_id_display = "***"
        ws.cell(row=row, column=7, value=payment_method_id_display).border = BORDER
        
        # Дата следующей попытки (если попытки < 3 и автопродление включено)
        next_attempt = "—"
        if auto_renewal_enabled and attempts and attempts < 3 and last_attempt_at:
            try:
                last_attempt_dt = datetime.fromisoformat(last_attempt_at.replace('Z', '+00:00'))
                if last_attempt_dt.tzinfo is None:
                    last_attempt_dt = pytz.utc.localize(last_attempt_dt)
                # Следующая попытка через 2 часа
                next_attempt_dt = last_attempt_dt + timedelta(hours=2)
                next_attempt = format_datetime(next_attempt_dt.isoformat())
            except:
                pass
        elif expires_at and auto_renewal_enabled:
            # Если подписка истекает и автопродление включено - первая попытка сразу после истечения
            next_attempt = format_datetime(expires_at)
        
        ws.cell(row=row, column=8, value=next_attempt).border = BORDER
        
        row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 25

def create_summary_sheet(wb, conn):
    """Создает сводный лист со статистикой"""
    ws = wb.create_sheet("Сводка", 0)  # Первый лист
    
    # Заголовок
    ws['A1'] = "СВОДНАЯ СТАТИСТИКА"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    
    cur = conn.cursor()
    
    # Общая статистика
    row = 3
    ws.cell(row=row, column=1, value="Параметр").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Значение").font = Font(bold=True)
    row += 1
    
    # Количество пользователей
    cur.execute("SELECT COUNT(*) FROM users")
    user_count = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Всего пользователей:")
    ws.cell(row=row, column=2, value=user_count)
    row += 1
    
    # Количество активных подписок
    cur.execute("""
        SELECT COUNT(*) FROM subscriptions 
        WHERE expires_at > datetime('now', 'utc')
    """)
    active_subs = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Активных подписок:")
    ws.cell(row=row, column=2, value=active_subs)
    row += 1
    
    # Истекших подписок
    cur.execute("""
        SELECT COUNT(*) FROM subscriptions 
        WHERE expires_at <= datetime('now', 'utc')
    """)
    expired_subs = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Истекших подписок:")
    ws.cell(row=row, column=2, value=expired_subs)
    row += 1
    
    # Подписок истекает в течение 24 часов
    cur.execute("""
        SELECT COUNT(*) FROM subscriptions 
        WHERE expires_at > datetime('now', 'utc')
        AND expires_at <= datetime('now', '+1 day', 'utc')
    """)
    expiring_24h = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Истекает в течение 24 часов:")
    ws.cell(row=row, column=2, value=expiring_24h)
    row += 1
    
    # Подписок истекает в течение 7 дней
    cur.execute("""
        SELECT COUNT(*) FROM subscriptions 
        WHERE expires_at > datetime('now', 'utc')
        AND expires_at <= datetime('now', '+7 days', 'utc')
    """)
    expiring_7d = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Истекает в течение 7 дней:")
    ws.cell(row=row, column=2, value=expiring_7d)
    row += 1
    
    # Количество успешных платежей
    cur.execute("SELECT COUNT(*) FROM payments WHERE status = 'succeeded'")
    success_payments = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Успешных платежей:")
    ws.cell(row=row, column=2, value=success_payments)
    row += 1
    
    # Платежи бонусной недели
    cur.execute("""
        SELECT COUNT(*) FROM payments 
        WHERE status = 'succeeded'
        AND created_at >= ? AND created_at <= ?
    """, (BONUS_WEEK_START_DATE.isoformat(), BONUS_WEEK_END_DATE.isoformat()))
    bonus_payments = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Платежей бонусной недели:")
    ws.cell(row=row, column=2, value=bonus_payments)
    row += 1
    
    # Общая сумма (успешные платежи * стоимость подписки)
    # Учитываем бонусную неделю
    bonus_amount = bonus_payments * float(BONUS_WEEK_PRICE_RUB)
    production_payments = success_payments - bonus_payments
    production_amount = production_payments * float(PRODUCTION_PRICE_RUB)
    total_amount = bonus_amount + production_amount
    ws.cell(row=row, column=1, value="Общая сумма (руб):")
    ws.cell(row=row, column=2, value=f"{total_amount:.2f}")
    row += 1
    
    # Средняя сумма платежа
    if success_payments > 0:
        avg_amount = total_amount / success_payments
        ws.cell(row=row, column=1, value="Средняя сумма платежа (руб):")
        ws.cell(row=row, column=2, value=f"{avg_amount:.2f}")
        row += 1
    
    # Платежи в ожидании
    cur.execute("SELECT COUNT(*) FROM payments WHERE status = 'pending'")
    pending_payments = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Платежей в ожидании:")
    ws.cell(row=row, column=2, value=pending_payments)
    row += 1
    
    # Отмененные платежи
    cur.execute("SELECT COUNT(*) FROM payments WHERE status = 'canceled'")
    canceled_payments = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Отмененных платежей:")
    ws.cell(row=row, column=2, value=canceled_payments)
    row += 1
    
    # Конверсия (успешные / все)
    cur.execute("SELECT COUNT(*) FROM payments")
    all_payments = cur.fetchone()[0]
    if all_payments > 0:
        conversion = (success_payments / all_payments) * 100
        ws.cell(row=row, column=1, value="Конверсия (%):")
        ws.cell(row=row, column=2, value=f"{conversion:.1f}%")
        row += 1
    
    # Пользователи с автопродлением
    cur.execute("SELECT COUNT(*) FROM subscriptions WHERE auto_renewal_enabled = 1")
    auto_renewal_count = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Пользователей с автопродлением:")
    ws.cell(row=row, column=2, value=auto_renewal_count)
    row += 1
    
    # Пользователи с неудачными попытками автопродления (3 попытки)
    cur.execute("SELECT COUNT(*) FROM subscriptions WHERE auto_renewal_attempts >= 3")
    failed_auto_renewal = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Неудачных автопродлений (3 попытки):")
    ws.cell(row=row, column=2, value=failed_auto_renewal)
    row += 1
    
    # Пользователи с формой заполнена
    cur.execute("SELECT COUNT(*) FROM users WHERE form_filled = 1")
    forms_filled = cur.fetchone()[0]
    ws.cell(row=row, column=1, value="Пользователей с заполненной формой:")
    ws.cell(row=row, column=2, value=forms_filled)
    row += 1
    
    # Средняя длительность активных подписок
    cur.execute("""
        SELECT AVG((julianday(expires_at) - julianday(starts_at))) 
        FROM subscriptions 
        WHERE expires_at > datetime('now', 'utc') AND starts_at IS NOT NULL
    """)
    avg_duration = cur.fetchone()[0]
    if avg_duration:
        ws.cell(row=row, column=1, value="Средняя длительность активных подписок (дн.):")
        ws.cell(row=row, column=2, value=f"{avg_duration:.1f}")
        row += 1
    
    # Среднее время обработки платежа
    cur.execute("""
        SELECT AVG((julianday(pp.processed_at) - julianday(p.created_at)) * 24 * 60)
        FROM payments p
        JOIN processed_payments pp ON p.payment_id = pp.payment_id
        WHERE p.status = 'succeeded'
    """)
    avg_processing = cur.fetchone()[0]
    if avg_processing:
        ws.cell(row=row, column=1, value="Среднее время обработки платежа (мин.):")
        ws.cell(row=row, column=2, value=f"{avg_processing:.1f}")
        row += 1
    
    # Дата генерации отчета
    row += 1
    ws.cell(row=row, column=1, value="Дата генерации отчета:")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%d.%m.%Y %H:%M:%S"))
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

def main():
    """Главная функция"""
    if not os.path.exists(DB_PATH):
        print(f"❌ Ошибка: База данных не найдена: {DB_PATH}")
        sys.exit(1)
    
    print(f"📊 Генерация Excel отчета из базы данных: {DB_PATH}")
    
    try:
        conn = sqlite3.connect(DB_PATH)
        
        # Создаем рабочую книгу
        wb = Workbook()
        
        # Создаем листы
        print("  ✓ Создание листа 'Сводка'...")
        create_summary_sheet(wb, conn)
        
        print("  ✓ Создание листа 'Пользователи'...")
        create_users_sheet(wb, conn)
        
        print("  ✓ Создание листа 'Платежи'...")
        create_payments_sheet(wb, conn)
        
        print("  ✓ Создание листа 'Подписки'...")
        create_subscriptions_sheet(wb, conn)
        
        print("  ✓ Создание листа 'Пригласительные ссылки'...")
        create_invite_links_sheet(wb, conn)
        
        print("  ✓ Создание листа 'Заполненные формы'...")
        create_forms_sheet(wb, conn)
        
        print("  ✓ Создание листа 'Автопродление'...")
        create_auto_renewal_sheet(wb, conn)
        
        # Сохраняем файл во временную директорию
        import tempfile
        temp_dir = tempfile.gettempdir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bot_report_{timestamp}.xlsx"
        file_path = os.path.join(temp_dir, filename)
        wb.save(file_path)
        
        conn.close()
        
        print(f"\n✅ Отчет успешно создан: {filename}")
        print(f"   Полный путь: {file_path}")
        
        # Возвращаем путь к файлу для использования в других скриптах
        return file_path
        
    except Exception as e:
        print(f"❌ Ошибка при создании отчета: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
